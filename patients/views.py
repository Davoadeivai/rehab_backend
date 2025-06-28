from django.contrib.auth import authenticate
from rest_framework import viewsets, status, mixins
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.authtoken.models import Token
from django.http import HttpResponse, JsonResponse
from .models import Patient, Notification
from .medication_models import (
    Service, ServiceTransaction, Medication, MedicationType, Prescription, MedicationDistribution, Payment, 
    DrugInventory, DrugAppointment, MedicationAdministration, DrugQuota, DrugReceipt
)
from openpyxl import Workbook
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import get_template
from django.conf import settings    
from django.utils import timezone
from django.utils.html import strip_tags
from django.contrib.auth.models import User
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.db.models.functions import ExtractMonth, ExtractYear
import json
from django.contrib.auth.tokens import default_token_generator
from django.core.mail import send_mail
from django.db.models import Count, Sum, Q, F, FloatField
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import jdatetime
from .utils import format_jalali_date, format_jalali_full_date, format_number
from django.views.generic import ListView, CreateView, DetailView, UpdateView
from django.urls import reverse_lazy, reverse
from rest_framework import generics
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import (JalaliDateField, PatientForm, PaymentForm, PrescriptionForm, MedicationDistributionForm,
                    UserProfileForm, UserSettingsForm, ContactForm, SupportForm, FeedbackForm, MedicationForm, MedicationAdministrationForm, ServiceTransactionForm)
from django.core.cache import cache
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.views.decorators.vary import vary_on_cookie
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
import json

from .serializers import (
    PatientSerializer,
    RegisterSerializer,
    LoginSerializer,
    PasswordResetSerializer,
    PasswordResetConfirmSerializer,
    MedicationTypeSerializer,
    PrescriptionSerializer,
    MedicationDistributionSerializer,
    PaymentSerializer,
)
from .services.medication_summary import monthly_medication_summary
import xlwt
import io
import logging

logger = logging.getLogger(__name__)

# -------------------------------
# Authentication API (Register & Login)
# -------------------------------

# بخش احراز هویت (Authentication)
# این بخش شامل توابع مربوط به مدیریت کاربران و احراز هویت آنها در سیستم است.
class AuthViewSet(viewsets.GenericViewSet):
    """
    مجموعه ویوهای مربوط به احراز هویت کاربران

    این کلاس شامل عملیات‌های زیر است:
    - ورود کاربران
    - ثبت‌نام کاربران جدید
    - بازیابی رمز عبور
    - تایید بازیابی رمز عبور
    """
    permission_classes = [AllowAny]
    serializer_class = RegisterSerializer

    # این تابع برای ورود کاربران به سیستم استفاده می‌شود.
    # این بخشی از API احراز هویت است و اطلاعات ورود کاربر را اعتبارسنجی می‌کند.
    @action(detail=False, methods=['post'], url_path='login', url_name='login')
    def login(self, request):
        """
        ورود کاربران به سیستم

        این تابع درخواست ورود کاربر را پردازش کرده و در صورت صحت اطلاعات، توکن دسترسی را برمی‌گرداند.

        پارامترها:
            request: شامل نام کاربری و رمز عبور

        برمی‌گرداند:
            Response: توکن دسترسی در صورت موفقیت، یا پیام خطا در صورت شکست
        """
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = authenticate(
                username=serializer.validated_data['username'],
                password=serializer.validated_data['password']
            )
            if user:
                token, _ = Token.objects.get_or_create(user=user)
                return Response({'token': token.key}, status=status.HTTP_200_OK)
            return Response({'error': 'نام کاربری یا رمز عبور اشتباه است'}, status=status.HTTP_401_UNAUTHORIZED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # این تابع برای ثبت‌نام کاربران جدید در سیستم استفاده می‌شود.
    # این بخشی از API احراز هویت است و اطلاعات ثبت‌نام کاربر را اعتبارسنجی و ذخیره می‌کند.
    # این تابع برای ثبت نام کاربران جدید در سیستم استفاده می‌شود.
    # اطلاعات کاربر جدید را اعتبارسنجی کرده و در صورت موفقیت، کاربر را ایجاد می‌کند.
    @action(detail=False, methods=['post'], url_path='register', url_name='register')
    def register(self, request):
        # این تابع برای ثبت نام کاربران جدید در سیستم استفاده می‌شود.
        # اطلاعات کاربر جدید را اعتبارسنجی کرده و در صورت موفقیت، کاربر را ایجاد می‌کند.
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.save()
        return Response({
            "user": RegisterSerializer(user, context=self.get_serializer_context()).data,
            "message": "User Created Successfully. Now perform Login to get Access Token",
        }, status=status.HTTP_201_CREATED)

    # این تابع برای درخواست بازنشانی رمز عبور کاربر استفاده می‌شود.
    # این بخشی از API احراز هویت است و ایمیل بازنشانی رمز عبور را ارسال می‌کند.
    @action(detail=False, methods=['post'], url_path='password-reset', url_name='password_reset')
    def password_reset(self, request):
        serializer = PasswordResetSerializer(data=request.data)
        if serializer.is_valid():
            email = serializer.validated_data.get('email')
            reset_url = serializer.validated_data.get('reset_url', '#')  # Ensure reset_url is provided by your serializer
            from django.core.mail import send_mail
            send_mail(
                'بازیابی رمز عبور',
                f'برای بازیابی رمز عبور خود روی لینک زیر کلیک کنید:\n\n{reset_url}',
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            return Response({"detail": "ایمیل بازیابی رمز عبور ارسال شد."}, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# این تابع برای تایید بازنشانی رمز عبور کاربر استفاده می‌شود.
# پس از اعتبارسنجی توکن و UID، رمز عبور جدید را برای کاربر تنظیم می‌کند.
@action(detail=False, methods=['post'], url_path='password-reset-confirm', url_name='password_reset_confirm')
def password_reset_confirm(self, request):
    serializer = PasswordResetConfirmSerializer(data=request.data)
    if serializer.is_valid():
        try:
            uid = force_str(urlsafe_base64_decode(serializer.validated_data['uidb64']))
            user = User.objects.get(pk=uid)
        except (TypeError, ValueError, OverflowError, User.DoesNotExist):
            return Response({"detail": "لینک نامعتبر است."}, status=status.HTTP_400_BAD_REQUEST)

        if default_token_generator.check_token(user, serializer.validated_data['token']):
            user.set_password(serializer.validated_data['password'])
            user.save()
            return Response({"detail": "رمز عبور با موفقیت تغییر کرد."})
            return Response({"detail": "توکن نامعتبر است."}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@login_required
def service_list(request):
    services = Service.objects.all()
    return render(request, 'patients/service_list.html', {'services': services})

@login_required
def service_transaction_create(request):
    if request.method == 'POST':
        form = ServiceTransactionForm(request.POST)
        if form.is_valid():
            transaction = form.save()
            messages.success(request, 'تراکنش با موفقیت ثبت شد.')
            # Check if the service is paid and redirect to payment form
            if transaction.service.price > 0:
                return redirect(f"{reverse('payment_create')}?patient_id={transaction.patient.id}&transaction_id={transaction.id}")
            return redirect('patient_detail', pk=transaction.patient.pk)
    else:
        form = ServiceTransactionForm()
    return render(request, 'patients/service_transaction_form.html', {'form': form})

class PatientViewSet(viewsets.ModelViewSet):
    """
    مجموعه ویوهای مربوط به مدیریت بیماران

    این کلاس شامل تمام عملیات‌های CRUD برای بیماران و همچنین عملیات‌های اضافی زیر است:
    - جستجوی پیشرفته بیماران
    - گزارش‌گیری جامع
    - آمار و ارقام
    - لیست بیماران فعال

    Attributes:
        queryset: کوئری پایه برای دریافت تمام بیماران
        serializer_class: کلاس سریالایزر برای تبدیل مدل به JSON
        permission_classes: کلاس‌های مجوز دسترسی
    """
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer
    permission_classes = []

    # این تابع لیست تمام بیماران را بازیابی می‌کند. 
    # از کش به مدت ۱۵ دقیقه برای بهبود عملکرد استفاده می‌کند.
    @method_decorator(cache_page(60 * 15))  # Cache for 15 minutes
    @method_decorator(vary_on_cookie)
    def list(self, request, *args, **kwargs):
        """
        لیست تمام بیماران با کش ۱۵ دقیقه‌ای
        """
        return super().list(request, *args, **kwargs)

    # این تابع برای بازیابی لیست بیماران با قابلیت فیلتر و جستجو استفاده می‌شود.
    # فیلترها شامل جنسیت، وضعیت تاهل، تحصیلات، نوع درمان و نوع دارو می‌شوند.
    # همچنین امکان جستجو بر اساس نام، نام خانوادگی، کد ملی و شماره پرونده وجود دارد.
    # فیلترهای تاریخ پذیرش و وضعیت درمان (فعال/تکمیل شده) نیز اعمال می‌شوند.
    def get_queryset(self):
        """
        بازیابی لیست بیماران با قابلیت فیلتر و جستجو

        Returns:
            QuerySet: لیست فیلتر شده بیماران با تمام روابط مورد نیاز
        """
        queryset = Patient.objects.prefetch_related(
            'prescription_set',
            'payment_set'
        ).all()
        
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(national_code__icontains=search) |
                Q(file_number__icontains=search)
            )

        filters = {
            'gender': 'gender',
            'marital_status': 'marital_status',
            'education': 'education',
            'treatment_type': 'treatment_type',
            'drug_type': 'drug_type',
        }
        
        for param, field in filters.items():
            value = self.request.query_params.get(param, None)
            if value:
                queryset = queryset.filter(**{field: value})

        admission_after = self.request.query_params.get('admission_after', None)
        admission_before = self.request.query_params.get('admission_before', None)
        
        if admission_after:
            queryset = queryset.filter(admission_date__gte=admission_after)
        if admission_before:
            queryset = queryset.filter(admission_date__lte=admission_before)

        treatment_status = self.request.query_params.get('treatment_status', None)
        if treatment_status:
            if treatment_status == 'active':
                queryset = queryset.filter(treatment_withdrawal_date__isnull=True)
            elif treatment_status == 'completed':
                queryset = queryset.filter(treatment_withdrawal_date__isnull=False)

        sort_by = self.request.query_params.get('sort_by', '-admission_date')
        if sort_by:
            queryset = queryset.order_by(sort_by)

        return queryset

    # این تابع برای ایجاد یک بیمار جدید در سیستم استفاده می‌شود.
    # پس از اعتبارسنجی اطلاعات، بیمار را ذخیره کرده و یک پیام موفقیت‌آمیز به همراه اطلاعات بیمار برمی‌گرداند.
    # همچنین اطلاعات مربوط به ایجاد بیمار را در لاگ ثبت می‌کند.
    def create(self, request, *args, **kwargs):
        """
        ایجاد یک بیمار جدید در سیستم
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        logger.info(
            'Patient created successfully - File Number: %s, Name: %s %s',
            serializer.data.get('file_number'),
            serializer.data.get('first_name'),
            serializer.data.get('last_name')
        )
        return Response(
            {"detail": "بیمار با موفقیت ثبت شد", "data": serializer.data},
            status=status.HTTP_201_CREATED,
            headers=headers
        )

    # این تابع برای به‌روزرسانی اطلاعات یک بیمار موجود استفاده می‌شود.
    # پس از اعتبارسنجی اطلاعات، بیمار را به‌روزرسانی کرده و یک پیام موفقیت‌آمیز به همراه اطلاعات به‌روز شده بیمار برمی‌گرداند.
    # همچنین اطلاعات مربوط به به‌روزرسانی بیمار را در لاگ ثبت می‌کند.
    def update(self, request, *args, **kwargs):
        """
        به‌روزرسانی اطلاعات یک بیمار موجود
        """
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        logger.info(
            'Patient updated successfully - File Number: %s, Name: %s %s',
            instance.file_number,
            instance.first_name,
            instance.last_name
        )
        return Response(
            {"detail": "اطلاعات بیمار با موفقیت به‌روز شد", "data": serializer.data}
        )

    # این تابع برای حذف یک بیمار از سیستم استفاده می‌شود.
    # پس از حذف، یک پیام موفقیت‌آمیز برمی‌گرداند و اطلاعات مربوط به حذف بیمار را در لاگ ثبت می‌کند.
    def destroy(self, request, *args, **kwargs):
        """
        حذف بیمار
        """
        instance = self.get_object()
        logger.warning(
            'Patient deleted - File Number: %s, Name: %s %s',
            instance.file_number,
            instance.first_name,
            instance.last_name
        )
        self.perform_destroy(instance)
        return Response(
            {"detail": "بیمار با موفقیت حذف شد"},
            status=status.HTTP_204_NO_CONTENT
        )

    # این تابع جزئیات کامل یک بیمار خاص را برمی‌گرداند.
    # از طریق `pk` (کلید اصلی) بیمار مورد نظر را شناسایی کرده و اطلاعات کامل آن را نمایش می‌دهد.
    @action(detail=True, methods=['get'])
    def full_details(self, request, pk=None):
        """
        نمایش جزئیات کامل یک بیمار
        """
        patient = self.get_object()
        serializer = self.get_serializer(patient)
        return Response(serializer.data)

    # این تابع لیست بیمارانی را برمی‌گرداند که هنوز در حال درمان هستند.
    # بیمارانی که تاریخ پایان درمان آنها مشخص نشده باشد، به عنوان بیماران فعال در نظر گرفته می‌شوند.
    @action(detail=False, methods=['get'])
    def active_patients(self, request):
        """
        لیست بیمارانی که هنوز در حال درمان هستند
        """
        active_patients = self.get_queryset().filter(treatment_withdrawal_date__isnull=True)
        serializer = self.get_serializer(active_patients, many=True)
        return Response(serializer.data)

    # این تابع آمار پیشرفته‌ای از بیماران ارائه می‌دهد.
    # شامل آمار کلی (تعداد کل، فعال، تکمیل شده)، آمار جنسیتی، آمار نوع درمان و آمار نوع مواد مصرفی است.
    # همچنین میانگین مدت درمان برای بیمارانی که درمانشان تکمیل شده را محاسبه می‌کند.
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """
        آمار پیشرفته بیماران
        """
        queryset = self.get_queryset()
        current_date = jdatetime.datetime.now().date()
        
        # آمار کلی
        total_patients = queryset.count()
        active_patients = queryset.filter(treatment_withdrawal_date__isnull=True).count()
        completed_treatments = queryset.filter(treatment_withdrawal_date__isnull=False).count()
        
        # آمار جنسیتی
        gender_stats = queryset.values('gender').annotate(count=Count('gender'))
        
        # آمار نوع درمان
        treatment_stats = queryset.values('treatment_type').annotate(count=Count('treatment_type'))
        
        # آمار نوع مواد مصرفی
        drug_stats = queryset.values('drug_type').annotate(count=Count('drug_type'))
        
        # میانگین مدت درمان برای درمان‌های تکمیل شده
        completed_treatments_qs = queryset.filter(
            treatment_withdrawal_date__isnull=False,
            admission_date__isnull=False
        )
        
        avg_treatment_duration = 0
        if completed_treatments_qs.exists():
            total_days = sum(
                (patient.treatment_withdrawal_date - patient.admission_date).days
                for patient in completed_treatments_qs
            )
            avg_treatment_duration = total_days / completed_treatments_qs.count()

        return Response({
            'total_statistics': {
                'total_patients': total_patients,
                'active_patients': active_patients,
                'completed_treatments': completed_treatments,
                'avg_treatment_duration': round(avg_treatment_duration, 1)
            },
            'gender_statistics': gender_stats,
            'treatment_type_statistics': treatment_stats,
            'drug_type_statistics': drug_stats,
        })

    # این تابع یک گزارش جامع از تمام فعالیت‌های یک بیمار خاص ارائه می‌دهد.
    # شامل اطلاعات شخصی، نسخه‌های دارویی، توزیع داروها، پرداخت‌ها، آمار کلی (نسخه‌شده، توزیع‌شده، باقی‌مانده، پرداخت‌ها)،
    # مدت درمان، آمار پرداخت‌ها بر اساس نوع و آمار ماهانه پرداخت‌ها است.
    # از کش به مدت ۵ دقیقه برای بهبود عملکرد استفاده می‌کند.
    @action(detail=True, methods=['get'])
    @method_decorator(cache_page(60 * 5))  # Cache for 5 minutes
    def comprehensive_report(self, request, pk=None):
        """
        گزارش جامع از تمام فعالیت‌های یک بیمار با کش ۵ دقیقه‌ای
        """
        cache_key = f'patient_report_{pk}'
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
        
        patient = self.get_object()
        
        # اطلاعات شخصی بیمار
        patient_data = self.get_serializer(patient).data
        
        # نسخه‌های دارویی
        prescriptions = Prescription.objects.filter(patient=patient).order_by('-start_date')
        prescription_data = PrescriptionSerializer(prescriptions, many=True).data
        
        # توزیع داروها
        distributions = MedicationDistribution.objects.filter(
            prescription__patient=patient
        ).order_by('-distribution_date')
        distribution_data = MedicationDistributionSerializer(distributions, many=True).data
        
        # پرداخت‌ها
        payments = Payment.objects.filter(patient=patient).order_by('-payment_date')
        payment_data = PaymentSerializer(payments, many=True).data
        
        # آمار کلی
        total_prescribed = prescriptions.aggregate(
            total=Sum('total_prescribed')
        )['total'] or 0
        
        total_distributed = distributions.aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        total_payments = payments.aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        # محاسبه مدت درمان
        if patient.admission_date:
            if patient.treatment_withdrawal_date:
                treatment_duration = (patient.treatment_withdrawal_date - patient.admission_date).days
                status = "اتمام درمان"
            else:
                current_date = jdatetime.datetime.now().date()
                treatment_duration = (current_date - patient.admission_date).days
                status = "در حال درمان"
        else:
            treatment_duration = 0
            status = "نامشخص"
        
        # آمار پرداخت‌ها بر اساس نوع
        payment_stats = payments.values('payment_period').annotate(
            total=Sum('amount'),
            count=Count('id')
        )
        
        # تبدیل تاریخ‌ها به فرمت فارسی
        jalali_dates = {
            'admission_date': format_jalali_full_date(patient.admission_date),
            'treatment_withdrawal_date': format_jalali_full_date(patient.treatment_withdrawal_date),
            'last_prescription': format_jalali_full_date(prescriptions.first().start_date) if prescriptions.exists() else '',
            'last_distribution': format_jalali_full_date(distributions.first().distribution_date) if distributions.exists() else '',
            'last_payment': format_jalali_full_date(payments.first().payment_date) if payments.exists() else '',
        }
        
        # آمار ماهانه پرداخت‌ها
        monthly_payments = payments.extra(
            select={'month': "EXTRACT(month FROM payment_date)"}
        ).values('month').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('month')
        
        # تبدیل شماره ماه به نام فارسی ماه
        persian_months = {
            1: 'فروردین', 2: 'اردیبهشت', 3: 'خرداد',
            4: 'تیر', 5: 'مرداد', 6: 'شهریور',
            7: 'مهر', 8: 'آبان', 9: 'آذر',
            10: 'دی', 11: 'بهمن', 12: 'اسفند'
        }
        
        monthly_payment_stats = [
            {
                'month': persian_months.get(item['month'], str(item['month'])),
                'total': item['total'],
                'count': item['count'],
                'total_display': "{:,}".format(item['total'])
            }
            for item in monthly_payments
        ]
        
        response_data = {
            'patient_info': patient_data,
            'treatment_status': {
                'status': status,
                'duration_days': treatment_duration,
                'important_dates': jalali_dates
            },
            'medication_summary': {
                'total_prescribed': total_prescribed,
                'total_distributed': total_distributed,
                'remaining': total_prescribed - total_distributed,
                'prescriptions_count': prescriptions.count(),
                'distributions_count': distributions.count()
            },
            'financial_summary': {
                'total_payments': total_payments,
                'total_payments_display': "{:,}".format(total_payments),
                'payment_stats': payment_stats,
                'monthly_stats': monthly_payment_stats,
                'payments_count': payments.count()
            },
            'prescriptions': prescription_data,
            'distributions': distribution_data,
            'payments': payment_data
        }
        
        cache.set(cache_key, response_data, timeout=60 * 5)  # Cache for 5 minutes
        return Response(response_data)

@login_required
def export_to_excel(request):
    """
    صدور گزارش اکسل از اطلاعات بیماران

    این تابع یک فایل اکسل شامل اطلاعات زیر تولید می‌کند:
    - اطلاعات شخصی بیماران
    - اطلاعات درمانی
    - تاریخ‌های مهم
    - وضعیت پرداخت‌ها

    برمی‌گرداند:
        HttpResponse: فایل اکسل آماده دانلود
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "گزارش بیماران"

    # تنظیم هدر ستون‌ها
    headers = [
        'شماره پرونده', 'نام', 'نام خانوادگی', 'کد ملی', 'تاریخ تولد',
        'جنسیت', 'وضعیت تأهل', 'تحصیلات', 'نوع ماده مصرفی',
        'نوع درمان', 'تاریخ پذیرش', 'تاریخ ترک درمان'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # اضافه کردن داده‌ها
    patients = Patient.objects.all()
    for row, patient in enumerate(patients, 2):
        ws.cell(row=row, column=1).value = patient.file_number
        ws.cell(row=row, column=2).value = patient.first_name
        ws.cell(row=row, column=3).value = patient.last_name
        ws.cell(row=row, column=4).value = patient.national_code
        ws.cell(row=row, column=5).value = format_jalali_date(patient.date_birth)
        ws.cell(row=row, column=6).value = patient.get_gender_display()
        ws.cell(row=row, column=7).value = patient.get_marital_status_display()
        ws.cell(row=row, column=8).value = patient.get_education_display()
        ws.cell(row=row, column=9).value = patient.get_drug_type_display()
        ws.cell(row=row, column=10).value = patient.get_treatment_type_display()
        ws.cell(row=row, column=11).value = format_jalali_date(patient.admission_date)
        ws.cell(row=row, column=12).value = format_jalali_date(patient.treatment_withdrawal_date)

    # تنظیم عرض ستون‌ها
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # ایجاد پاسخ HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=patients_report.xlsx'
    
    wb.save(response)
    return response

# این تابع برای صدور لیست بیماران به صورت فایل PDF استفاده می‌شود.
# اطلاعات بیماران را از پایگاه داده بازیابی کرده و با استفاده از یک قالب HTML، آن را به PDF تبدیل می‌کند.
@login_required
def export_to_pdf(request):
    template_path = 'patients/patient_list_pdf.html'
    context = {'patients': Patient.objects.all()}
    
    # Create a Django response object, and specify content_type as pdf
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="patients.pdf"'
    
    # Find the template and render it
    template = get_template(template_path)
    html = template.render(context)
    
    # Create a PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    # If error then show some funny view
    if pisa_status.err:
        return HttpResponse('خطا در ایجاد PDF')
    return response

# مجموعه ویوهای مربوط به مدیریت انواع داروها
# این کلاس شامل عملیات‌های CRUD برای انواع داروها است.
class MedicationTypeViewSet(viewsets.ModelViewSet):
    queryset = MedicationType.objects.all()
    serializer_class = MedicationTypeSerializer
    permission_classes = []

    # این تابع کوئری‌ست پایه را برای بازیابی انواع داروها برمی‌گرداند.
    def get_queryset(self):
        queryset = MedicationType.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(name__icontains=search)
        return queryset

# مجموعه ویوهای مربوط به مدیریت نسخه‌ها
# این کلاس شامل عملیات‌های CRUD برای نسخه‌ها و همچنین محاسبه داروی باقی‌مانده است.
class PrescriptionViewSet(viewsets.ModelViewSet):
    queryset = Prescription.objects.all()
    serializer_class = PrescriptionSerializer
    permission_classes = []

    # این تابع کوئری‌ست پایه را برای بازیابی نسخه‌ها با قابلیت فیلتر و جستجو بر اساس بیمار، نوع دارو و تاریخ برمی‌گرداند.
    def get_queryset(self):
        """
        بازیابی لیست نسخه‌ها با روابط از پیش بارگذاری شده

        Returns:
            QuerySet: لیست نسخه‌ها با تمام روابط مورد نیاز
        """
        queryset = Prescription.objects.select_related(
            'patient',
            'medication_type'
        ).prefetch_related(
            'medicationdistribution_set'
        ).all()
        
        # فیلتر بر اساس بیمار
        patient_id = self.request.query_params.get('patient', None)
        if patient_id:
            queryset = queryset.filter(patient__file_number=patient_id)
        
        # فیلتر بر اساس نوع دارو
        medication_type = self.request.query_params.get('medication_type', None)
        if medication_type:
            queryset = queryset.filter(medication_type_id=medication_type)
        
        # فیلتر بر اساس تاریخ
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        
        return queryset

    # این تابع مقدار داروی باقی‌مانده برای یک نسخه خاص را محاسبه و برمی‌گرداند.
    @action(detail=True, methods=['get'])
    def remaining_medication(self, request, pk=None):
        """محاسبه داروی باقی‌مانده برای یک نسخه"""
        prescription = self.get_object()
        total_distributed = MedicationDistribution.objects.filter(
            prescription=prescription
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        remaining = prescription.total_prescribed - total_distributed
        
        return Response({
            'total_prescribed': prescription.total_prescribed,
            'total_distributed': total_distributed,
            'remaining': remaining
        })

class MedicationDistributionViewSet(viewsets.ModelViewSet):
    """
    مجموعه ویوهای مربوط به توزیع دارو

    این کلاس شامل عملیات‌های زیر است:
    - ثبت توزیع دارو
    - مشاهده لیست توزیع‌ها
    - به‌روزرسانی و حذف توزیع‌ها
    - محاسبه مقدار باقی‌مانده دارو

    Attributes:
        queryset: کوئری پایه برای دریافت تمام توزیع‌های دارو
        serializer_class: کلاس سریالایزر برای تبدیل مدل به JSON
        permission_classes: کلاس‌های مجوز دسترسی
    """
    queryset = MedicationDistribution.objects.all()
    serializer_class = MedicationDistributionSerializer
    permission_classes = []

    # این تابع کوئری‌ست پایه را برای بازیابی توزیع داروها با قابلیت فیلتر و جستجو بر اساس نسخه، بیمار و تاریخ برمی‌گرداند.
    def get_queryset(self):
        """
        بازیابی لیست توزیع داروها با روابط از پیش بارگذاری شده

        Returns:
            QuerySet: لیست توزیع داروها با تمام روابط مورد نیاز
        """
        queryset = MedicationDistribution.objects.select_related(
            'prescription',
            'prescription__patient',
            'prescription__medication_type'
        ).all()
        
        # فیلتر بر اساس نسخه
        prescription_id = self.request.query_params.get('prescription', None)
        if prescription_id:
            queryset = queryset.filter(prescription_id=prescription_id)
        
        # فیلتر بر اساس بیمار
        patient_id = self.request.query_params.get('patient', None)
        if patient_id:
            queryset = queryset.filter(prescription__patient__file_number=patient_id)
        
        # فیلتر بر اساس تاریخ
        date = self.request.query_params.get('date', None)
        if date:
            queryset = queryset.filter(distribution_date=date)
        
        return queryset

    # این تابع عملیات ایجاد توزیع دارو را انجام می‌دهد و اعتبارسنجی می‌کند که مقدار توزیع شده از مقدار تجویز شده بیشتر نباشد.
    def perform_create(self, serializer):
        """
        اجرای عملیات ایجاد توزیع دارو با اعتبارسنجی مقدار باقی‌مانده
        """
        prescription = serializer.validated_data['prescription']
        amount = serializer.validated_data['amount']
        
        total_distributed = MedicationDistribution.objects.filter(
            prescription=prescription
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        remaining = prescription.total_prescribed - (total_distributed + amount)
        
        if remaining < 0:
            logger.error(
                'Medication distribution failed - Prescription ID: %s, Amount: %s, Remaining: %s',
                prescription.id,
                amount,
                remaining
            )
            raise serializers.ValidationError(
                "مقدار توزیع شده بیشتر از مقدار تجویز شده است."
            )
        
        distribution = serializer.save(remaining=remaining)
        logger.info(
            'Medication distributed successfully - Prescription ID: %s, Amount: %s, Remaining: %s',
            prescription.id,
            amount,
            remaining
        )

class PaymentViewSet(viewsets.ModelViewSet):
    """
    مجموعه ویوهای مربوط به پرداخت‌ها

    این کلاس شامل عملیات‌های زیر است:
    - ثبت پرداخت جدید
    - مشاهده لیست پرداخت‌ها
    - گزارش‌گیری از پرداخت‌ها
    - محاسبه آمار پرداخت‌ها

    Attributes:
        queryset: کوئری پایه برای دریافت تمام پرداخت‌ها
        serializer_class: کلاس سریالایزر برای تبدیل مدل به JSON
        permission_classes: کلاس‌های مجوز دسترسی
    """
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    permission_classes = []

    # این تابع کوئری‌ست پایه را برای بازیابی پرداخت‌ها با قابلیت فیلتر و جستجو بر اساس بیمار، نوع پرداخت و تاریخ برمی‌گرداند.
    @method_decorator(cache_page(60 * 15))  # Cache for 15 minutes
    def get_queryset(self):
        """
        بازیابی لیست پرداخت‌ها با روابط از پیش بارگذاری شده

        Returns:
            QuerySet: لیست پرداخت‌ها با تمام روابط مورد نیاز
        """
        queryset = Payment.objects.select_related('patient').all()
        
        # فیلتر بر اساس بیمار
        patient_id = self.request.query_params.get('patient', None)
        if patient_id:
            queryset = queryset.filter(patient__file_number=patient_id)
        
        # فیلتر بر اساس نوع پرداخت
        payment_type = self.request.query_params.get('payment_type', None)
        if payment_type:
            queryset = queryset.filter(payment_period=payment_type)
        
        # فیلتر بر اساس تاریخ
        start_date = self.request.query_params.get('start_date', None)
        end_date = self.request.query_params.get('end_date', None)
        
        if start_date:
            queryset = queryset.filter(payment_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(payment_date__lte=end_date)
        
        return queryset

    # این تابع یک گزارش خلاصه از پرداخت‌ها را با قابلیت فیلتر بر اساس تاریخ، دوره (روزانه، هفتگی، ماهانه) و بیمار ارائه می‌دهد.
    # نتایج به مدت ۱۵ دقیقه کش می‌شوند.
    @method_decorator(cache_page(60 * 15))  # Cache for 15 minutes
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        گزارش خلاصه پرداخت‌ها با کش ۱۵ دقیقه‌ای
        """
        cache_key = 'payment_summary'
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
            
        # پارامترهای فیلتر
        start_date = request.query_params.get('start_date', None)
        end_date = request.query_params.get('end_date', None)
        period = request.query_params.get('period', 'daily')  # daily, weekly, monthly
        patient_id = request.query_params.get('patient', None)
        
        # پایه کوئری
        queryset = self.get_queryset()
        
        if start_date:
            queryset = queryset.filter(payment_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(payment_date__lte=end_date)
        if patient_id:
            queryset = queryset.filter(patient__file_number=patient_id)
        
        # تجمیع بر اساس دوره
        if period == 'daily':
            payments = queryset.values('payment_date').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('payment_date')
        elif period == 'weekly':
            payments = queryset.extra(select={'week': "EXTRACT(week FROM payment_date)"}).values(
                'week'
            ).annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('week')
        else:  # monthly
            payments = queryset.extra(select={'month': "EXTRACT(month FROM payment_date)"}).values(
                'month'
            ).annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('month')
        
        # تجمیع بر اساس نوع پرداخت
        payment_types = queryset.values('payment_type').annotate(
            total=Sum('amount'),
            count=Count('id')
        )
        
        response_data = {
            'summary_by_period': payments,
            'summary_by_type': payment_types,
            'total_amount': queryset.aggregate(total=Sum('amount'))['total'],
            'total_count': queryset.count()
        }
        
        cache.set(cache_key, response_data, timeout=60 * 15)  # Cache for 15 minutes
        return Response(response_data)

    # این تابع یک پرداخت جدید را ایجاد می‌کند و اطلاعات مربوط به آن را ثبت می‌کند.
    def create(self, request, *args, **kwargs):
        """
        ایجاد پرداخت جدید
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        logger.info(
            'Payment created successfully - Patient: %s, Amount: %s, Type: %s',
            serializer.data.get('patient'),
            serializer.data.get('amount'),
            serializer.data.get('payment_type')
        )
        return Response(
            serializer.data,
            status=status.HTTP_201_CREATED,
            headers=headers
        )

# Template Views
@login_required
# این تابع لیست تمام بیماران را با قابلیت جستجو و صفحه‌بندی بازیابی کرده و نمایش می‌دهد.
# برای بهینه‌سازی عملکرد، نتایج صفحه‌بندی شده و کوئری‌ها بهینه هستند.
def patient_list(request):
    # دریافت کوئری جستجو از پارامتر GET
    query = request.GET.get('q', '')
    
    # شروع با تمام بیماران و مرتب‌سازی بر اساس جدیدترین
    patient_queryset = Patient.objects.all().order_by('-created_at')

    # اگر کوئری جستجو وجود داشت، لیست را فیلتر کن
    if query:
        patient_queryset = patient_queryset.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(national_code__icontains=query) |
            Q(file_number__icontains=query)
        )

    # ایجاد یک شیء Paginator با ۲۵ آیتم در هر صفحه
    paginator = Paginator(patient_queryset, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # ارسال شیء صفحه‌بندی شده و کوئری به قالب
    return render(request, 'patients/patient_list.html', {
        'page_obj': page_obj,
        'query': query
    })

# این تابع برای جستجوی بیماران بر اساس نام، نام خانوادگی، کد ملی و وضعیت درمان استفاده می‌شود.
# نتایج جستجو در یک صفحه HTML نمایش داده می‌شوند.
@login_required
def patient_search(request):
    query = request.GET.get('q', '')
    status = request.GET.get('status', '')
    
    patients = Patient.objects.all()
    
    if query:
        patients = patients.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(national_id__icontains=query)
        )
    
    if status:
        if status == 'active':
            patients = patients.filter(treatment_withdrawal_date__isnull=True)
        elif status == 'inactive':
            patients = patients.filter(treatment_withdrawal_date__isnull=False)
    
    # Pagination
    paginator = Paginator(patients, 10)  # Show 10 patients per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'patients': page_obj,
        'query': query,
        'status': status,
    }
    
    return render(request, 'patients/search.html', context)

# این تابع برای ثبت بیمار جدید استفاده می‌شود.
# اگر درخواست از نوع POST باشد، اطلاعات فرم اعتبارسنجی شده و بیمار جدید ذخیره می‌شود.
# در غیر این صورت، فرم خالی برای ثبت بیمار نمایش داده می‌شود.
import sys

@login_required
def patient_create(request):
    print('patient_create view called', file=sys.stderr)
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            print('Form is valid', file=sys.stderr)
            patient = form.save(commit=False)
            national_code = form.cleaned_data.get('national_code')
            if not patient.file_number and national_code:
                base_file_number = national_code[-4:]
                new_file_number = base_file_number
                counter = 1
                while Patient.objects.filter(file_number=new_file_number).exists():
                    new_file_number = f"{base_file_number}-{counter}"
                    counter += 1
                patient.file_number = new_file_number
            print(f'Generated file_number: {patient.file_number}', file=sys.stderr)
            try:
                patient.save()
                print(f'Patient saved successfully, pk={patient.pk}', file=sys.stderr)
                messages.success(request, 'بیمار جدید با موفقیت ثبت شد.')
                # Redirect to patient list after successful save
                return redirect('patients:patient_list')
            except Exception as e:
                print(f'Error on save: {e}', file=sys.stderr)
                form.add_error(None, f'خطا در ذخیره بیمار: {e}')
        else:
            print(f'Form is invalid: {form.errors}', file=sys.stderr)
    else:
        form = PatientForm()
    return render(request, 'patients/patient_form.html', {'form': form, 'title': 'ثبت بیمار جدید'})

# این تابع جزئیات یک بیمار خاص را بر اساس کلید اصلی (pk) بازیابی کرده و در قالب یک صفحه HTML نمایش می‌دهد.
@login_required
def patient_detail(request, pk):
    patient = get_object_or_404(Patient, pk=pk)

    # Financial calculations
    service_transactions = ServiceTransaction.objects.filter(patient=patient).order_by('-date')
    payments = Payment.objects.filter(patient=patient).order_by('-payment_date')

    total_cost = service_transactions.aggregate(
        total=Sum(F('service__price') * F('quantity'), output_field=FloatField()))['total'] or 0.0

    total_paid = payments.aggregate(total=Sum('amount'))['total'] or 0.0

    balance = total_cost - total_paid

    # Medication history
    medication_distributions = MedicationDistribution.objects.filter(patient=patient).order_by('-distribution_date')

    context = {
        'patient': patient,
        'service_transactions': service_transactions,
        'payments': payments,
        'total_cost': total_cost,
        'total_paid': total_paid,
        'balance': balance,
        'medication_distributions': medication_distributions,
    }
    return render(request, 'patients/patient_detail.html', context)


@login_required
def dashboard(request):
    # Key statistics
    total_patients = Patient.objects.count()
    total_revenue = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    total_cost = ServiceTransaction.objects.aggregate(
        total=Sum(F('service__price') * F('quantity'), output_field=FloatField()))['total'] or 0.0

    # Recent activities
    recent_patients = Patient.objects.order_by('-admission_date')[:5]
    recent_payments = Payment.objects.order_by('-payment_date')[:5]
    recent_transactions = ServiceTransaction.objects.order_by('-date')[:5]

    context = {
        'total_patients': total_patients,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'recent_patients': recent_patients,
        'recent_payments': recent_payments,
        'recent_transactions': recent_transactions,
    }
    return render(request, 'patients/dashboard.html', context)

# این تابع برای ویرایش اطلاعات بیمار موجود استفاده می‌شود.
# اگر درخواست از نوع POST باشد، اطلاعات فرم اعتبارسنجی شده و اطلاعات بیمار به‌روزرسانی می‌شود.
# در غیر این صورت، فرم با اطلاعات فعلی بیمار برای ویرایش نمایش داده می‌شود.
@login_required
def patient_edit(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            patient = form.save()
            messages.success(request, 'اطلاعات بیمار با موفقیت به‌روزرسانی شد.')
            return redirect('patients:patient_detail', pk=patient.pk)
    else:
        form = PatientForm(instance=patient)
    return render(request, 'patients/patient_form.html', {'form': form, 'title': 'ویرایش اطلاعات بیمار'})

@login_required
def inventory_list(request):
    """نمایش لیست موجودی داروها"""
    inventory_items = DrugInventory.objects.select_related('medication').all()
    low_stock_items = inventory_items.filter(current_stock__lt=F('minimum_stock'))
    
    context = {
        'inventory_items': inventory_items,
        'low_stock_items': low_stock_items,
    }
    return render(request, 'patients/inventory_list.html', context)

# این تابع برای حذف یک بیمار استفاده می‌شود.
# اگر درخواست از نوع POST باشد، بیمار مورد نظر حذف شده و کاربر به لیست بیماران هدایت می‌شود.
# در غیر این صورت، صفحه تأیید حذف بیمار نمایش داده می‌شود.
@login_required
def patient_delete(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        patient.delete()
        messages.success(request, 'بیمار با موفقیت حذف شد.')
        return redirect('patients:patient_list')
    return render(request, 'patients/patient_confirm_delete.html', {'patient': patient})

# API Views
class PatientListCreateAPIView(generics.ListCreateAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer

class PatientRetrieveUpdateDestroyAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer

@login_required
def prescription_create(request):
    import sys
    print('prescription_create view called', file=sys.stderr)
    """
    ایجاد نسخه جدید برای بیمار

    این تابع امکان ثبت نسخه جدید با مشخصات زیر را فراهم می‌کند:
    - انتخاب بیمار
    - انتخاب نوع دارو
    - تعیین دوز روزانه
    - تعیین مدت درمان
    - تعیین تاریخ شروع و پایان
    - ثبت یادداشت‌های اضافی

    برمی‌گرداند:
        HttpResponse: صفحه فرم ثبت نسخه یا ریدایرکت به لیست نسخه‌ها
    """
    if request.method == 'POST':
        print('POST data:', request.POST, file=sys.stderr)
        form = PrescriptionForm(request.POST)
        if form.is_valid():
            print('Form is valid', file=sys.stderr)
            prescription = form.save()
            prescription = form.save()
            messages.success(request, 'نسخه با موفقیت ثبت شد. لطفاً اطلاعات پرداخت را تکمیل کنید.')
            # Redirect to the payment creation page with the prescription ID
            payment_create_url = reverse('patients:payment_create')
            return redirect(f'{payment_create_url}?prescription_id={prescription.id}')
        else:
            print(f'Form is invalid: {form.errors}', file=sys.stderr)
    else:
        form = PrescriptionForm()
    
    return render(request, 'patients/prescription_form.html', {
        'form': form,
        'title': 'ثبت نسخه جدید'
    })

@login_required
def payment_create(request):
    """
    ثبت پرداخت جدید برای یک نسخه یا به صورت عمومی.

    اگر `prescription_id` در URL وجود داشته باشد، فرم با اطلاعات
    بیمار و مبلغ محاسبه‌شده از روی نسخه پر می‌شود.
    در غیر این صورت، یک فرم خالی برای ثبت پرداخت عمومی نمایش داده می‌شود.
    """
    prescription = None
    initial_data = {}
    prescription_id = request.GET.get('prescription_id')

    if prescription_id:
        prescription = get_object_or_404(Prescription, id=prescription_id)
        initial_data = {
            'prescription': prescription,
            'patient': prescription.patient,
        }
        # فرض: هزینه نسخه بر اساس یک سرویس همنام با نوع دارو محاسبه می‌شود
        try:
            service = Service.objects.get(name=prescription.medication_type.name, service_type='drug')
            # فرض: هزینه برابر است با قیمت واحد سرویس ضربدر مقدار کل تجویز شده
            cost = service.unit_price * prescription.total_prescribed
            initial_data['amount'] = cost.quantize(Decimal('0.01'))
        except Service.DoesNotExist:
            messages.warning(request, f"سرویس قیمت‌گذاری برای داروی '{prescription.medication_type.name}' یافت نشد. لطفاً مبلغ را دستی وارد کنید.")
        except Exception as e:
            messages.error(request, f"خطا در محاسبه هزینه: {str(e)}")

    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            # اگر پرداخت برای یک نسخه خاص است، بیمار را دستی تنظیم می‌کنیم
            # چون فیلد غیرفعال در POST ارسال نمی‌شود
            if prescription:
                payment.patient = prescription.patient
            payment.save()
            form.save_m2m()  # برای ذخیره فیلدهای ManyToMany
            messages.success(request, 'پرداخت با موفقیت ثبت شد.')
            # TODO: در مرحله بعد به درگاه پرداخت هدایت شود
            return redirect('patients:payment_list')
    else:
        form = PaymentForm(initial=initial_data)
        # اگر پرداخت برای یک نسخه خاص است، فیلد بیمار غیرفعال شود
        if prescription:
            form.fields['patient'].disabled = True

    return render(request, 'patients/payment_form.html', {
        'form': form,
        'title': 'ثبت پرداخت جدید',
    })

@login_required
def prescription_list(request):
    prescriptions = Prescription.objects.all().order_by('-created_at')
    return render(request, 'patients/prescription_list.html', {'prescriptions': prescriptions})

@login_required
def distribution_list(request):
    distributions = MedicationDistribution.objects.all().order_by('-distribution_date')
    return render(request, 'patients/distribution_list.html', {'distributions': distributions})

@login_required
def distribution_create(request):
    prescription_id = request.GET.get('prescription')
    initial_data = {'prescription': prescription_id} if prescription_id else {}
    
    if request.method == 'POST':
        form = MedicationDistributionForm(request.POST)
        if form.is_valid():
            distribution = form.save()
            messages.success(request, 'توزیع دارو با موفقیت ثبت شد.')
            if distribution.prescription.patient_id:
                return redirect('patients:patient_detail', pk=distribution.prescription.patient_id)
            return redirect('patients:distribution_list')
    else:
        form = MedicationDistributionForm(initial=initial_data)
    
    return render(request, 'patients/distribution_form.html', {
        'form': form,
        'prescription_id': prescription_id,
        'title': 'ثبت توزیع دارو'
    })

@login_required
def payment_list(request):
    payments = Payment.objects.all().order_by('-payment_date')
    total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
    return render(request, 'patients/payment_list.html', {
        'payments': payments,
        'total_amount': total_amount
    })

@login_required
@cache_page(60 * 15)  # Cache for 15 minutes
def report_list(request):
    """
    نمایش لیست گزارش‌ها با کش ۱۵ دقیقه‌ای
    """
    cache_key = 'report_list'
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return render(request, 'patients/report_list.html', cached_data)
    
    # آمار کلی
    total_patients = Patient.objects.count()
    active_patients = Patient.objects.filter(treatment_withdrawal_date__isnull=True).count()
    completed_patients = Patient.objects.filter(treatment_withdrawal_date__isnull=False).count()
    
    # آمار پرداخت‌ها
    total_payments = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    payment_by_type = Payment.objects.values('payment_type').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    # آمار توزیع دارو
    total_distributions = MedicationDistribution.objects.count()
    medication_stats = MedicationDistribution.objects.values(
        'prescription__medication_type__name'
    ).annotate(
        total_amount=Sum('amount'),
        count=Count('id')
    )
    
    context = {
        'total_patients': total_patients,
        'active_patients': active_patients,
        'completed_patients': completed_patients,
        'total_payments': total_payments,
        'payment_by_type': payment_by_type,
        'total_distributions': total_distributions,
        'medication_stats': medication_stats,
    }
    
    cache.set(cache_key, context, timeout=60 * 15)  # Cache for 15 minutes
    return render(request, 'patients/report_list.html', context)

@login_required
def payment_detail(request, pk):
    payment = get_object_or_404(
        Payment.objects.select_related(
            'patient',
            'prescription__doctor'
        ).prefetch_related(
            'prescription__medications__medication_type'
        ),
        pk=pk
    )
    return render(request, 'patients/payment_detail.html', {'payment': payment})

@login_required
def payment_edit(request, pk):
    """ویرایش پرداخت"""
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            payment = form.save()
            messages.success(request, 'پرداخت با موفقیت به‌روزرسانی شد.')
            return redirect('patients:payment_list')
    else:
        form = PaymentForm(instance=payment)
    
    return render(request, 'patients/payment_form.html', {
        'form': form,
        'title': 'ویرایش پرداخت'
    })

@login_required
def payment_delete(request, pk):
    """حذف پرداخت"""
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        payment.delete()
        messages.success(request, 'پرداخت با موفقیت حذف شد.')
        return redirect('patients:payment_list')
    return render(request, 'patients/payment_confirm_delete.html', {'payment': payment})

class UpdateInventoryView(UpdateView):
    model = DrugInventory
    fields = ['current_stock', 'minimum_stock']
    template_name = 'patients/inventory_update.html'
    success_url = reverse_lazy('patients:inventory_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'موجودی دارو با موفقیت به‌روزرسانی شد.')
        return response

@login_required
def inventory_view(request):
    """نمایش لیست موجودی داروها"""
    inventory_items = DrugInventory.objects.select_related('medication_type').all()
    low_stock_items = inventory_items.filter(current_stock__lt=F('minimum_stock'))
    
    context = {
        'inventory_items': inventory_items,
        'low_stock_items': low_stock_items,
    }
    return render(request, 'patients/inventory_list.html', context)

# این تابع گزارش‌های مالی مربوط به پرداخت‌های بیماران را تهیه و نمایش می‌دهد.
# شامل مجموع پرداخت‌ها، پرداخت‌ها بر اساس دوره و پرداخت‌های ماهانه است.
@login_required
def financial_reports(request):
    payments = Payment.objects.all().order_by('-payment_date')
    total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
    
    payment_by_type = payments.values('transactions__service__service_type').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    monthly_payments = payments.extra(
        select={'month': "EXTRACT(month FROM payment_date)"}
    ).values('month').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('month')
    
    context = {
        'total_amount': total_amount,
        'payment_by_type': payment_by_type,
        'monthly_payments': monthly_payments,
    }
    return render(request, 'patients/financial_reports.html', context)

@login_required
def patient_reports(request):
    """آمار بیماران"""
    total_patients = Patient.objects.count()
    active_patients = Patient.objects.filter(treatment_withdrawal_date__isnull=True).count()
    completed_patients = Patient.objects.filter(treatment_withdrawal_date__isnull=False).count()
    
    # آمار جنسیتی
    gender_stats = Patient.objects.values('gender').annotate(count=Count('gender'))
    
    # آمار نوع درمان
    treatment_stats = Patient.objects.values('treatment_type').annotate(count=Count('treatment_type'))
    
    context = {
        'total_patients': total_patients,
        'active_patients': active_patients,
        'completed_patients': completed_patients,
        'gender_stats': gender_stats,
        'treatment_stats': treatment_stats,
    }
    return render(request, 'patients/patient_reports.html', context)

@login_required
def prescription_reports(request):
    """آمار نسخه‌ها"""
    prescriptions = Prescription.objects.all()
    total_prescriptions = prescriptions.count()
    
    # آمار نوع دارو
    medication_stats = prescriptions.values(
        'medication_type__name'
    ).annotate(
        count=Count('id'),
        total_amount=Sum('total_prescribed')
    )
    
    # آمار توزیع دارو
    distribution_stats = MedicationDistribution.objects.values(
        'prescription__medication_type__name'
    ).annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    )
    
    context = {
        'total_prescriptions': total_prescriptions,
        'medication_stats': medication_stats,
        'distribution_stats': distribution_stats,
    }
    return render(request, 'patients/prescription_reports.html', context)

@login_required
def profile(request):
    """نمایش و ویرایش پروفایل کاربر"""
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'پروفایل شما با موفقیت بروزرسانی شد.')
            return redirect('profile')
    else:
        form = UserProfileForm(instance=request.user)
    
    return render(request, 'patients/profile.html', {'form': form})

@login_required
def settings(request):
    """تنظیمات کاربر"""
    if request.method == 'POST':
        form = UserSettingsForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'تنظیمات شما با موفقیت بروزرسانی شد.')
            return redirect('settings')
    else:
        form = UserSettingsForm(instance=request.user)

@login_required
def home(request):
    """صفحه اصلی - داشبورد مدیریتی"""
    # آمار کلی
    total_patients = Patient.objects.count()
    active_patients = Patient.objects.filter(treatment_withdrawal_date__isnull=True).count()
    total_revenue = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0
    total_prescriptions = Prescription.objects.count()
    total_cost = ServiceTransaction.objects.aggregate(
        total=Sum(F('service__unit_price') * F('quantity'), output_field=FloatField()))['total'] or 0.0

    # آخرین فعالیت‌ها
    recent_patients = Patient.objects.order_by('-admission_date')[:5]
    recent_payments = Payment.objects.order_by('-payment_date')[:5]
    recent_prescriptions = Prescription.objects.order_by('-created_at')[:5]  # Assuming created_at exists
    recent_transactions = ServiceTransaction.objects.order_by('-date')[:5]

    context = {
        'total_patients': total_patients,
        'active_patients': active_patients,
        'total_payments': total_revenue, # Renamed for consistency
        'total_prescriptions': total_prescriptions,
        'total_cost': total_cost,
        'recent_patients': recent_patients,
        'recent_payments': recent_payments,
        'recent_prescriptions': recent_prescriptions,
        'recent_transactions': recent_transactions,
    }
    return render(request, 'patients/home.html', context)

@login_required
def contact(request):
    """صفحه تماس با ما"""
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'پیام شما با موفقیت ارسال شد.')
            return redirect('patients:contact')
    else:
        form = ContactForm()
    return render(request, 'patients/contact.html', {'form': form})

@login_required
def faq(request):
    """صفحه سوالات متداول"""
    faqs = FAQ.objects.all()
    return render(request, 'patients/faq.html', {'faqs': faqs})

@login_required
def docs(request):
    """صفحه مستندات سیستم"""
    docs = Documentation.objects.all()
    return render(request, 'patients/docs.html', {'docs': docs})

@login_required
def support(request):
    """صفحه پشتیبانی فنی"""
    if request.method == 'POST':
        form = SupportForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'درخواست پشتیبانی شما با موفقیت ثبت شد.')
            return redirect('patients:support')
    else:
        form = SupportForm()
    return render(request, 'patients/support.html', {'form': form})

@login_required
def feedback(request):
    """صفحه ارسال پیشنهادات"""
    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'پیشنهاد شما با موفقیت ثبت شد.')
            return redirect('patients:feedback')
    else:
        form = FeedbackForm()
    return render(request, 'patients/feedback.html', {'form': form})    

@login_required
def drug_appointment_calendar(request):
    patients = Patient.objects.all()
    return render(request, 'patients/drug_appointment_calendar.html', {'patients': patients})

@login_required
def drug_appointments_json(request):
    logger.info("drug_appointments_json view called.")
    appointments = DrugAppointment.objects.select_related('patient').all()
    events = []
    for appt in appointments:
        events.append({
            'id': appt.id,
            'title': f"{appt.patient.first_name} {appt.patient.last_name}",
            'start': appt.date.strftime('%Y-%m-%d'),
            'color': '#4e73df' if appt.is_paid else '#e74c3c',
            'extendedProps': {
                'amount': appt.amount,
                'is_paid': appt.is_paid,
                'patient_id': appt.patient.id,
            }
        })
    logger.info(f"Returning {len(events)} appointments.")
    return JsonResponse(events, safe=False)

# این تابع برای ایجاد یک نوبت دارویی جدید برای بیمار استفاده می‌شود.
# اطلاعات نوبت شامل شناسه بیمار، تاریخ (شمسی)، مبلغ و وضعیت پرداخت را دریافت می‌کند.
# تاریخ شمسی را به میلادی تبدیل کرده و نوبت دارویی را در پایگاه داده ذخیره می‌کند.
@csrf_exempt
@login_required
def create_drug_appointment(request):
    logger.info("create_drug_appointment view called.")
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            logger.info(f"Received data: {data}")
            patient_id = data.get('patient_id')
            date_str = data.get('date')  # مثال: '1404/03/17'
            amount = data.get('amount')
            is_paid = data.get('is_paid', False)
            patient = Patient.objects.get(id=patient_id)

            # تبدیل تاریخ شمسی به میلادی
            try:
                if '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts) == 3:
                        jy, jm, jd = map(int, parts)
                        gdate = jdatetime.date(jy, jm, jd).togregorian()
                    else:
                        return JsonResponse({'status': 'error', 'msg': 'فرمت تاریخ نامعتبر است.'}, status=400)
                else:
                    return JsonResponse({'status': 'error', 'msg': 'فرمت تاریخ نامعتبر است.'}, status=400)
            except Exception as e:
                logger.error(f"Error converting date {date_str}: {e}")
                return JsonResponse({'status': 'error', 'msg': f'خطا در تبدیل تاریخ: {e}'}, status=400)

            appt = DrugAppointment.objects.create(patient=patient, date=gdate, amount=amount, is_paid=is_paid)
            logger.info(f"DrugAppointment created with ID: {appt.id}")
            return JsonResponse({'status': 'ok', 'id': appt.id})
        except json.JSONDecodeError as e:
            logger.error(f"JSONDecodeError: {e}")
            return JsonResponse({'status': 'error', 'msg': 'Invalid JSON format.'}, status=400)
        except Exception as e:
            logger.error(f"An unexpected error occurred: {e}")
            return JsonResponse({'status': 'error', 'msg': 'An unexpected error occurred.'}, status=500)
    logger.info("create_drug_appointment received non-POST request.")
    return JsonResponse({'status': 'error', 'msg': 'Only POST requests are allowed.'}, status=400)

# ... (rest of the code remains the same)
# این تابع برای ثبت نام کاربران جدید از طریق فرم وب استفاده می‌شود.
# اگر درخواست POST باشد، فرم ثبت نام را اعتبارسنجی کرده و کاربر جدیدی ایجاد می‌کند.
# در غیر این صورت، فرم خالی ثبت نام را نمایش می‌دهد.
def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('patients:home')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

# این تابع جزئیات یک نسخه خاص را نمایش می‌دهد.
# شامل اطلاعات نسخه و لیست توزیع‌های دارویی مرتبط با آن نسخه است.
@login_required
def prescription_detail(request, pk):
    prescription = get_object_or_404(Prescription, pk=pk)
    distributions = prescription.medicationdistribution_set.all().order_by('-distribution_date')
    return render(request, 'patients/prescription_detail.html', {
        'prescription': prescription,
        'distributions': distributions
    })

@login_required
def prescription_update(request, pk):
    prescription = get_object_or_404(Prescription, pk=pk)
    if request.method == 'POST':
        form = PrescriptionForm(request.POST, instance=prescription)
        if form.is_valid():
            form.save()
            messages.success(request, 'نسخه با موفقیت بروزرسانی شد.')
            return redirect('patients:prescription_detail', pk=prescription.pk)
    else:
        form = PrescriptionForm(instance=prescription)
    
    return render(request, 'patients/prescription_form.html', {
        'form': form,
        'title': 'ویرایش نسخه',
        'prescription': prescription
    })

@login_required
def prescription_delete(request, pk):
    prescription = get_object_or_404(Prescription, pk=pk)
    if request.method == 'POST':
        prescription.delete()
        messages.success(request, 'نسخه با موفقیت حذف شد.')
        return redirect('patients:prescription_list')
    return render(request, 'patients/prescription_confirm_delete.html', {'prescription': prescription})

@login_required
def medication_list(request):
    # This is a placeholder for the medication list view
    # You would typically fetch medication data here
    medications = [] # Replace with actual medication data
    return render(request, 'patients/medication_list.html', {'medications': medications})

@login_required
def payment_create(request):
    initial_data = {}
    patient_id = request.GET.get('patient_id')
    transaction_id = request.GET.get('transaction_id')

    if patient_id:
        initial_data['patient'] = patient_id
    if transaction_id:
        initial_data['transactions'] = [transaction_id]

    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'پرداخت با موفقیت ثبت شد.')
            return redirect('patients:payment_list')
    else:
        form = PaymentForm(initial=initial_data)
        
    return render(request, 'patients/payment_form.html', {'form': form})

@login_required
def medication_create(request):
    if request.method == 'POST':
        form = MedicationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'دارو با موفقیت ثبت شد.')
            return redirect('patients:medication_list')
    else:
        form = MedicationForm()
    return render(request, 'patients/medication_form.html', {'form': form})

@login_required
def medication_administration_create(request):
    if request.method == 'POST':
        form = MedicationAdministrationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'تجویز دارو با موفقیت ثبت شد.')
            return redirect('patients:medication_administration_list') # Assuming a list view will exist
    else:
        form = MedicationAdministrationForm()
    return render(request, 'patients/medication_administration_form.html', {'form': form})

def get_notifications(request):
    notifications = Notification.objects.filter(is_read=False).order_by('-created_at')[:5]
    data = [{'title': n.title, 'message': n.message, 'created_at': n.created_at.strftime('%Y-%m-%d %H:%M:%S')} for n in notifications]
    return JsonResponse(data, safe=False)


@login_required
def get_medication_details(request, medication_id):
    try:
        medication = get_object_or_404(MedicationType, pk=medication_id)
        data = {
            'default_dose': medication.default_dose,
            'unit': medication.unit
        }
        return JsonResponse(data)
    except MedicationType.DoesNotExist:
        return JsonResponse({'error': 'Medication not found'}, status=404)


def calculate_end_date(request):
    start_date_str = request.GET.get('start_date')
    duration_str = request.GET.get('duration')

    if not start_date_str or not duration_str:
        return JsonResponse({'error': 'Missing start_date or duration'}, status=400)

    try:
        # Convert Jalali start date to Gregorian
        start_year, start_month, start_day = map(int, start_date_str.split('/'))
        gregorian_start_date = JalaliToGregorian(start_year, start_month, start_day).get_gregorian_date()
        
        # Add duration
        duration_days = int(duration_str)
        gregorian_end_date = gregorian_start_date + timedelta(days=duration_days)
        
        # Convert Gregorian end date back to Jalali
        jalali_end_date = GregorianToJalali(gregorian_end_date.year, gregorian_end_date.month, gregorian_end_date.day)
        end_date_str = f"{jalali_end_date.jyear}/{jalali_end_date.jmonth:02d}/{jalali_end_date.jday:02d}"
        
        return JsonResponse({'end_date': end_date_str})

    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid date or duration format'}, status=400)

@login_required
def medication_administration_list(request):
    administrations = MedicationAdministration.objects.all().order_by('-administration_date')
    context = {
        'administrations': administrations
    }
    return render(request, 'patients/medication_administration_list.html', context)


@login_required
def medication_administration_delete(request, pk):
    administration = get_object_or_404(MedicationAdministration, pk=pk)
    administration.delete()
    messages.success(request, 'تجویز دارو با موفقیت حذف شد.')
    return redirect('patients:medication_administration_list')

from django.contrib import messages
from django.db.models import Sum, F, FloatField
from .pos_service import send_payment_to_pos
from django.http import JsonResponse

@login_required
def payment_list(request):
    payments = Payment.objects.all().order_by('-payment_date')
    total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
    return render(request, 'patients/payment_list.html', {
        'payments': payments,
        'total_amount': total_amount
    })

@login_required
def initiate_pos_payment(request, pk):
    """
    Initiates a payment process via the POS terminal.

    This view is called via AJAX from the payment detail page.
    It finds the payment record, sends the request to the POS service,
    updates the payment status based on the response, and returns a JSON response.
    """
    if request.method == 'POST':
        try:
            payment = Payment.objects.get(pk=pk)
        except Payment.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'پرداخت یافت نشد.'}, status=404)

        # ارسال درخواست به سرویس پوز
        pos_response = send_payment_to_pos(payment.id, float(payment.amount))

        # به‌روزرسانی رکورد پرداخت بر اساس پاسخ دریافتی
        payment.status = pos_response.get('status', 'failed')
        payment.pos_transaction_id = pos_response.get('transaction_id')
        payment.pos_data = pos_response.get('pos_data')
        payment.save()

        return JsonResponse({
            'status': payment.status,
            'message': pos_response.get('message', 'خطای نامشخص'),
            'transaction_id': payment.pos_transaction_id
        })

    return JsonResponse({'status': 'error', 'message': 'درخواست نامعتبر.'}, status=400)